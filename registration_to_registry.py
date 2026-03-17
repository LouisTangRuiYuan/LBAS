
from __future__ import annotations

import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


REGION_NAME = "South West"
REGISTRY_SHEET_NAME = "South West"

# File keywords so the script can auto-find the two Excel files
REGISTRY_KEYWORDS = ["registry", "database"]
FORM_KEYWORDS = ["registration", "responses"]

# --- Mapping assumptions ---
# Registry columns:
# A ID (already prefilled in template, e.g. SW001)
# B Actual ID                <- SSA-UK Membership ID
# C NAME                     <- Full Name
# D Email                    <- E-mail (fallback Email Address)
# E Contact Number           <- WhatsApp-reachable number
# F Status                   <- "Student" or "Working / Graduate"
# G Number of dependents     <- numeric count
# H City                     <- City of Residence / "Others" text
# I University               <- University / Institution
# J Confirmation email sent? <- left blank

FORM_HEADERS_REQUIRED = {
    "Full Name": "full_name",
    "SSA-UK Membership ID": "membership_id",
    "E-mail": "email",
    "Email Address": "email_address",
    "UK/EU Phone Number ": "uk_phone",
    "Malaysia Phone Number (if available) ": "my_phone",
    "Which Phone Number will be reachable through WhatsApp? ": "whatsapp_pref",
    "Which region are you from?": "region",
    "City of Residence in UK and Ireland": "city",
    "If you answered 'Others' above, please specify which city": "other_city",
    "Current Status in UK/EU": "current_status",
    "Employment Status": "employment_status",
    "University / Institution ": "university",
    "How many dependents are you planning to bring to the event?": "dependents_count",
}


def find_excel_file(folder: Path, keywords: List[str], exclude_name: Optional[str] = None) -> Optional[Path]:
    candidates = []
    for ext in ("*.xlsx", "*.xlsm"):
        candidates.extend(folder.glob(ext))

    exclude_name = exclude_name.lower() if exclude_name else None

    scored: List[Tuple[int, Path]] = []
    for path in candidates:
        name = path.name.lower()
        if exclude_name and name == exclude_name:
            continue
        score = sum(1 for kw in keywords if kw in name)
        if score > 0:
            scored.append((score, path))

    if not scored:
        return None

    scored.sort(key=lambda x: (-x[0], x[1].name))
    return scored[0][1]


def normalize_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_email(value) -> str:
    return normalize_text(value).lower()


def normalize_name(value) -> str:
    return " ".join(normalize_text(value).upper().split())


def parse_whatsapp_number(row: Dict[str, object]) -> str:
    pref = normalize_text(row.get("whatsapp_pref"))
    uk_phone = normalize_text(row.get("uk_phone"))
    my_phone = normalize_text(row.get("my_phone"))

    pref_lower = pref.lower()
    if "uk number" in pref_lower and uk_phone:
        return uk_phone
    if "malaysian number" in pref_lower and my_phone:
        return my_phone
    return uk_phone or my_phone


def parse_status(row: Dict[str, object]) -> str:
    current_status = normalize_text(row.get("current_status")).lower()
    employment_status = normalize_text(row.get("employment_status"))

    if "student" in current_status:
        return "Student"

    # Any non-student case is grouped into the registry's broader bucket
    if employment_status or current_status:
        return "Working / Graduate"

    return ""


def parse_city(row: Dict[str, object]) -> str:
    city = normalize_text(row.get("city"))
    other_city = normalize_text(row.get("other_city"))

    if city.lower().startswith("other") and other_city:
        return other_city
    return city or other_city


def parse_dependents_count(value) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except Exception:
        return 0


def row_is_empty(ws, row_idx: int, start_col: int = 2, end_col: int = 10) -> bool:
    for col in range(start_col, end_col + 1):
        if ws.cell(row_idx, col).value not in (None, ""):
            return False
    return True


def get_form_header_map(ws) -> Dict[str, int]:
    header_map = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header is not None:
            header_map[str(header)] = col
    return header_map


def build_form_row(ws, row_idx: int, header_map: Dict[str, int]) -> Dict[str, object]:
    row = {}
    for excel_header, key in FORM_HEADERS_REQUIRED.items():
        col_idx = header_map.get(excel_header)
        row[key] = ws.cell(row_idx, col_idx).value if col_idx else None
    return row


def existing_keys_in_registry(ws) -> set:
    keys = set()
    for row_idx in range(2, ws.max_row + 1):
        name = normalize_name(ws.cell(row_idx, 3).value)
        email = normalize_email(ws.cell(row_idx, 4).value)
        actual_id = normalize_text(ws.cell(row_idx, 2).value)

        if email:
            keys.add(("email", email))
        if name:
            keys.add(("name", name))
        if actual_id:
            keys.add(("actual_id", actual_id))
    return keys


def choose_duplicate_keys(record: Dict[str, object]) -> List[Tuple[str, str]]:
    keys = []
    email = normalize_email(record["email"])
    name = normalize_name(record["name"])
    actual_id = normalize_text(record["actual_id"])

    if email:
        keys.append(("email", email))
    if actual_id:
        keys.append(("actual_id", actual_id))
    if name:
        keys.append(("name", name))
    return keys


def next_available_row(ws) -> int:
    for row_idx in range(2, ws.max_row + 1):
        if row_is_empty(ws, row_idx, 2, 10):
            return row_idx
    return ws.max_row + 1


def copy_style_from_previous_row(ws, row_idx: int) -> None:
    # Optional formatting preservation if a new row needs to be added past the template area
    if row_idx <= 2:
        return

    from copy import copy

    prev_row = row_idx - 1
    for col in range(1, 11):
        src = ws.cell(prev_row, col)
        dst = ws.cell(row_idx, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)


def make_record(form_row: Dict[str, object]) -> Dict[str, object]:
    email = normalize_text(form_row.get("email")) or normalize_text(form_row.get("email_address"))
    return {
        "actual_id": normalize_text(form_row.get("membership_id")),
        "name": normalize_text(form_row.get("full_name")),
        "email": email,
        "contact": parse_whatsapp_number(form_row),
        "status": parse_status(form_row),
        "dependents": parse_dependents_count(form_row.get("dependents_count")),
        "city": parse_city(form_row),
        "university": normalize_text(form_row.get("university")),
    }


def main() -> int:
    folder = Path(__file__).resolve().parent

    registry_path = find_excel_file(folder, REGISTRY_KEYWORDS)
    form_path = find_excel_file(folder, FORM_KEYWORDS, exclude_name=registry_path.name if registry_path else None)

    if registry_path is None:
        print("Could not find the registry workbook in this folder.")
        print("Expected filename to include words like: registry, database")
        return 1

    if form_path is None:
        print("Could not find the registration form workbook in this folder.")
        print("Expected filename to include words like: registration, responses")
        return 1

    keep_vba_registry = registry_path.suffix.lower() == ".xlsm"
    keep_vba_form = form_path.suffix.lower() == ".xlsm"

    registry_wb = load_workbook(registry_path, keep_vba=keep_vba_registry)
    form_wb = load_workbook(form_path, data_only=False, keep_vba=keep_vba_form)

    if REGISTRY_SHEET_NAME not in registry_wb.sheetnames:
        print(f'Sheet "{REGISTRY_SHEET_NAME}" not found in registry workbook.')
        return 1

    form_ws = form_wb[form_wb.sheetnames[0]]
    registry_ws = registry_wb[REGISTRY_SHEET_NAME]

    header_map = get_form_header_map(form_ws)

    missing_headers = [h for h in FORM_HEADERS_REQUIRED if h not in header_map]
    if missing_headers:
        print("The registration form is missing expected columns:")
        for h in missing_headers:
            print(f" - {h}")
        return 1

    registry_keys = existing_keys_in_registry(registry_ws)

    added_count = 0
    skipped_duplicates = 0
    south_west_found = 0

    for row_idx in range(2, form_ws.max_row + 1):
        form_row = build_form_row(form_ws, row_idx, header_map)
        region = normalize_text(form_row.get("region"))

        if region != REGION_NAME:
            continue

        south_west_found += 1
        record = make_record(form_row)

        duplicate = False
        for key in choose_duplicate_keys(record):
            if key in registry_keys:
                duplicate = True
                break

        if duplicate:
            skipped_duplicates += 1
            continue

        target_row = next_available_row(registry_ws)
        if target_row > registry_ws.max_row:
            copy_style_from_previous_row(registry_ws, target_row)

        # If column A has no ID formula/value for this row, continue the numbering pattern.
        if registry_ws.cell(target_row, 1).value in (None, ""):
            registry_ws.cell(target_row, 1).value = f"SW{target_row - 1:03d}"

        registry_ws.cell(target_row, 2).value = record["actual_id"] or None
        registry_ws.cell(target_row, 3).value = record["name"] or None
        registry_ws.cell(target_row, 4).value = record["email"] or None
        registry_ws.cell(target_row, 5).value = record["contact"] or None
        registry_ws.cell(target_row, 6).value = record["status"] or None
        registry_ws.cell(target_row, 7).value = record["dependents"]
        registry_ws.cell(target_row, 8).value = record["city"] or None
        registry_ws.cell(target_row, 9).value = record["university"] or None
        # Column J intentionally left blank

        # Update duplicate keys so reruns within same session don't duplicate
        for key in choose_duplicate_keys(record):
            registry_keys.add(key)

        added_count += 1

    registry_wb.save(registry_path)

    print(f"Registry file   : {registry_path.name}")
    print(f"Form file       : {form_path.name}")
    print(f"Region checked  : {REGION_NAME}")
    print(f"Rows found      : {south_west_found}")
    print(f"Rows added      : {added_count}")
    print(f"Skipped existing: {skipped_duplicates}")
    print("Done.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
